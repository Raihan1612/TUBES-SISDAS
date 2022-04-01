from openpyxl import load_workbook
import random
import xlsxwriter
import math
import numpy as np
import pandas as pd

def Node():
    #Membuat File Excel
    workbook = xlsxwriter.Workbook("Node.xlsx")
    #Membuat Sheet Excel
    worksheet = workbook.add_worksheet('Node')
    
    #Mengisi Label Kolom dan Baris
    worksheet.write('A1', 'No')
    worksheet.write('B1', 'Label')
    worksheet.write('C1', 'X')
    worksheet.write('D1', 'Y')
    
    #Mengisi hasil x dan y
    for row in range(n):
        no = row+1
        label = no
        convert_label = 'N'+str(label) 
        x = 5* random.randint(1,20)
        y = 5* random.randint(1,20)

        worksheet.write('A'+str(no+1), no)
        worksheet.write('B'+str(no+1), convert_label)
        worksheet.write('C'+str(no+1), x)
        worksheet.write('D'+str(no+1), y)
        
        #print(no, convert_label, x, y)

    
    workbook.close()

def Tetangga():
    #Menginisialisasi File Tetangga.xlsx
    workbook = xlsxwriter.Workbook("Tetangga.xlsx")
    worksheet = workbook.add_worksheet('Tetangga')
    
    huruf = 'A'

    for row in range(n):
        no_row = row+1
        convert_angka = 'N'+str(no_row)
        worksheet.write('A'+str(no_row+1), convert_angka) 
    
    for coloumn in range(n):
        no_col = coloumn+1
        huruf = chr(ord(huruf)+1)
        convert_huruf = huruf + '1'
        worksheet.write(convert_huruf, 'N'+str(no_col))

    #Men-generate ketetanggaan
    #Variabel yang dibutuhkan
    arr = [0] * n 
    next_coloumn = 0
    n_loop = n
    next_huruf = 0
    node_arr = 0
    #For Loop Untuk Row
    for row in range(2,n+2,1):
        temp_coloumn = 0
        next_huruf2 = next_huruf
        row_copy = 0
        node_itr = node_arr
        #For Loop Untuk Column
        for coloumn in range(65,65+n_loop,1):
            huruf = chr(coloumn+next_coloumn+1)
            huruf2 = chr(coloumn-next_huruf2+1) #Penjelasan Penggunaan Huruf2 Ada Di Bawah
            next_huruf2 += 1
            convert_huruf = huruf + str(row)
            convert_huruf2 = huruf2 + str(row+row_copy)
            hasil_random = random.choice([0,1])
            #Pengecekan Kondisi Apakah Hasil Random adalah 1
            if hasil_random == 1:
                temp_coloumn += 1
                #Pengecekan Kondisi Apakah Tiap Node Memiliki Ketetanggaan Kurang Dari 5
                if temp_coloumn <= 5 - arr[node_arr] and arr[node_arr]<5:
                    worksheet.write(convert_huruf, hasil_random)
                    worksheet.write(convert_huruf2, hasil_random)
                    arr[node_itr] += 1
                else:
                    worksheet.write(convert_huruf, 0)
                    worksheet.write(convert_huruf2, 0)
            else:
                worksheet.write(convert_huruf, hasil_random)
                worksheet.write(convert_huruf2, hasil_random)
            node_itr += 1
            row_copy += 1
        next_huruf -= 1
        next_coloumn += 1
        n_loop-=1
        node_arr+=1
    '''
    asumsi input = 1
    hasil convert_huruf tanpa convert_huruf2
    n 1 2 3 4 5
    1 1 1 1 1 1
    2   1 1 1 1
    3     1 1 1
    4       1 1
    5         1
    hasil convert_huruf dengan convert_huruf2
    n 1 2 3 4 5
    1 1 1 1 1 1
    2 1 1 1 1 1
    3 1 1 1 1 1
    4 1 1 1 1 1
    5 1 1 1 1 1
    '''
    workbook.close()

def Bobot():
    #menambahkan Bobot.xlsx
    workbook = xlsxwriter.Workbook("Bobot.xlsx")
    worksheet = workbook.add_worksheet('Bobot')
    
    huruf = 'A'

    #membuat baris pertama
    for row in range(n):
        no_row = row+1
        convert_angka = 'N'+str(no_row)
        worksheet.write('A'+str(no_row+1), convert_angka) 
    
    #membuat kolom pertama
    for coloumn in range(n):
        no_col = coloumn+1
        huruf = chr(ord(huruf)+1)
        convert_huruf = huruf + '1'
        worksheet.write(convert_huruf, 'N'+str(no_col))

    #melakukan akses pada file Tetangga.xlsx
    tetangga_workbook = load_workbook(filename="Tetangga.xlsx")
    tetangga_workbook.sheetnames
    tetangga_sheet = tetangga_workbook.active

    next_coloumn = 0
    m = n
    next_huruf = 0
    z = 0
    for row in range(2,n+2,1):
        next_huruf2 = next_huruf
        x = 0
        y = z
        for coloumn in range(65,65+m,1):
            huruf = chr(coloumn+next_coloumn+1)
            huruf2 = chr(coloumn-next_huruf2+1)
            next_huruf2 += 1
            convert_huruf = huruf + str(row)
            convert_huruf2 = huruf2 + str(row+x)

            if tetangga_sheet[convert_huruf].value == 1:
                temp_angka1 = ord(huruf) - 65
                temp_angka2 = ord(huruf2) - 65
                temp_label1 = "N" + str(temp_angka1)
                temp_label2 = "N" + str(temp_angka2)

                if temp_label1 != temp_label2:
                    hasil = fungsi_jarak(temp_label1, temp_label2,temp_angka1,temp_angka2,n)
                    worksheet.write(convert_huruf,hasil)
                    worksheet.write(convert_huruf2,hasil)
                else:
                    worksheet.write(convert_huruf,0)
                    worksheet.write(convert_huruf2,0)   
            
            else:
                worksheet.write(convert_huruf,0)
                worksheet.write(convert_huruf2,0)
                
            y += 1
            x += 1
        next_huruf -= 1
        next_coloumn += 1
        m-=1
        z+=1
    workbook.close()

def fungsi_jarak(label1,label2,angka1,angka2,n): #Fungsi untuk menghitung jarak antara 2 titik
    #Akses file Node.xlsx
    node_workbook = load_workbook(filename="Node.xlsx")
    node_workbook.sheetnames
    node_sheet = node_workbook.active
    
    #Melakukan iterasi untuk cek kolom nilai x dan y dari suatu label
    for baris in range(2,n+2,1):
        temp_label = "B" + str(baris)
        if node_sheet[temp_label].value == label1:
            x2 = node_sheet["C" + str(angka1+1)].value
            y2 = node_sheet["D" + str(angka1+1)].value
        elif node_sheet[temp_label].value == label2:
            x1 = node_sheet["C" + str(angka2+1)].value
            y1 = node_sheet["D" + str(angka2+1)].value
    
    #rumus jarak euclidean
    hasil = math.sqrt((x2-x1)**2+(y2-y1)**2)
    return hasil
    
def arrbobot():
    Node = []
    bobot_workbook = load_workbook(filename="Bobot.xlsx")
    bobot_workbook.sheetnames
    bobot_workbook = bobot_workbook.active

    arr_bobot = np.zeros((n, n), dtype=int)
    #print(arr_bobot)
    
    for i in range(n):
        label = "N"+str(i+1)
        Node.append(label)
        i+=1


    for i in range(n):
        j = 0
        for j in range(i+1):
            huruf = chr(66+j)
            convert_huruf = huruf + str(i+2)
            arr_bobot[i][j] = bobot_workbook[convert_huruf].value 
            arr_bobot[j][i] = bobot_workbook[convert_huruf].value 
            j+=1
        #print(i,j)
    bobot = pd.DataFrame(arr_bobot, index=Node, columns=Node)
    #print(bobot)
    return bobot

def find_shortest_path(graph, starting_node, goal):
    visited = []
    queue = [[starting_node]]
    
    while queue:
        path = queue.pop(0)
        node = path[-1]
        if node not in visited:
            neighbours = []
            for edge in graph:
                if edge[0] == node:
                    neighbours.append(edge[1])
                elif edge[1] == node:
                    neighbours.append(edge[0])
            for neighbour in neighbours:
                new_path = list(path)
                new_path.append(neighbour)
                queue.append(new_path)
                
                if neighbour == goal:
                    return new_path
            
            visited.append(node)
            
    return []

if __name__== "__main__":
    
    #Menjalankan Input n
    n = int(input("Masukkan rentang 10..20 : "))
    while (n < 10) or (n > 20):
        print("Anda Salah memasukkan input")
        n = int(input("Masukkan rentang 10..20 : "))
    
    
    Node()
    Tetangga()
    Bobot()
    frame_bobot = arrbobot()
    G = frame_bobot.replace(0, pd.NA).stack().index.to_list()
    awal = input("Awal: ")
    akhir = input("Akhir: ")
    print(find_shortest_path(G,awal,akhir))
